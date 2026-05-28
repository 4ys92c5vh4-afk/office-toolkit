#!/usr/bin/env python3
"""
读取 Excel 文件 (.xlsx)
支持：数据预览、多 sheet 读取、统计分析
"""

import sys
from pathlib import Path


def read_xlsx_with_pandas(filepath):
    """使用 pandas 读取 Excel"""
    import pandas as pd

    # 读取所有 sheet
    all_sheets = pd.read_excel(filepath, sheet_name=None)

    print(f"📊 发现 {len(all_sheets)} 个 Sheet:")
    for name in all_sheets.keys():
        print(f"   - {name}")

    for sheet_name, df in all_sheets.items():
        print(f"\n{'='*60}")
        print(f"📑 Sheet: {sheet_name}")
        print(f"{'='*60}")
        print(f"   行数: {len(df)} | 列数: {len(df.columns)}")
        print(f"   列名: {list(df.columns)}")
        print(f"\n📋 前 10 行数据:")
        print(df.head(10).to_string())
        print(f"\n📈 统计信息:")
        print(df.describe().to_string())

    return all_sheets


def read_xlsx_with_openpyxl(filepath):
    """使用 openpyxl 读取（保留公式）"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=False)
    print(f"\n🔧 openpyxl 读取 (保留公式):")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   Sheet '{sheet_name}': {ws.max_row} 行 x {ws.max_column} 列")
        # 显示前 5 行
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
            values = [cell.value for cell in row]
            print(f"      {values}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 read_xlsx.py <文件路径>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"❌ 文件不存在: {filepath}")
        sys.exit(1)

    print(f"📊 读取 Excel 文件: {filepath}")
    print("=" * 60)

    read_xlsx_with_pandas(filepath)
    read_xlsx_with_openpyxl(filepath)
