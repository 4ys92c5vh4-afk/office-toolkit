#!/usr/bin/env python3
"""
创建 Excel 文件 (.xlsx)
使用 openpyxl 创建带公式、格式、图表的工作簿
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_workbook(output_path):
    """创建示例工作簿"""
    wb = Workbook()

    # ========== Sheet 1: 销售数据 ==========
    ws1 = wb.active
    ws1.title = "销售数据"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["月份", "产品A", "产品B", "产品C", "总计"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据
    months = ["1月", "2月", "3月", "4月", "5月", "6月"]
    data = [
        [120, 150, 80],
        [135, 160, 95],
        [140, 155, 110],
        [160, 170, 120],
        [155, 180, 115],
        [170, 190, 130],
    ]

    for row_idx, (month, values) in enumerate(zip(months, data), 2):
        ws1.cell(row=row_idx, column=1, value=month).border = thin_border
        for col_idx, val in enumerate(values, 2):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # 公式：总计
        total_cell = ws1.cell(row=row_idx, column=5)
        total_cell.value = f"=SUM(B{row_idx}:D{row_idx})"
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.font = Font(bold=True)

    # 合计行
    ws1.cell(row=8, column=1, value="合计").font = Font(bold=True)
    for col in range(2, 6):
        cell = ws1.cell(row=8, column=col)
        cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}7)"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = thin_border

    # 调整列宽
    ws1.column_dimensions['A'].width = 10
    for col in ['B', 'C', 'D', 'E']:
        ws1.column_dimensions[col].width = 12

    # ========== Sheet 2: 产品列表 ==========
    ws2 = wb.create_sheet("产品列表")
    products = [
        ["ID", "产品名称", "类别", "单价", "库存"],
        ["P001", "笔记本电脑", "电子产品", 5999, 50],
        ["P002", "无线鼠标", "配件", 129, 200],
        ["P003", "机械键盘", "配件", 399, 80],
        ["P004", "显示器", "电子产品", 1499, 30],
    ]
    for row_data in products:
        ws2.append(row_data)

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 10

    wb.save(output_path)
    print(f"✅ Excel 已创建: {output_path}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "output.xlsx"
    create_workbook(output)
